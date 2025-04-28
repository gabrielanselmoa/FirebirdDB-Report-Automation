import os
import fdb
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

DATABASE_PATH = os.getenv("DATABASE_PATH")
USER = os.getenv("USER")
PASSWORD = os.getenv("PASSWORD")
OUTPUT_PATH = os.getenv("OUTPUT_PATH", "C:\\Users\\Gabriel Anselmo\\Desktop\\dados_tratados_formatados.xlsx")

con = None
cur = None

try:
    # Conectar ao banco Firebird
    con = fdb.connect(dsn=DATABASE_PATH, user=USER, password=PASSWORD)
    cur = con.cursor()

    # Executar consulta
    cur.execute("SELECT * FROM contratos;")
    dados = cur.fetchall()

    # Obter nomes das colunas
    colunas = [desc[0] for desc in cur.description]

    # Criar DataFrame Pandas
    df = pd.DataFrame(dados, columns=colunas)

    if df.empty:
        print("Nenhum dado encontrado na tabela 'contratos'.")
    else:
        # Tratamento de dados
        df.dropna(inplace=True)

        # Criar Workbook no openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Tratados"

        # Adicionar o DataFrame à planilha
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:  # Linha 1 = Cabeçalho
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        # Verificar se o arquivo já existe
        if os.path.exists(OUTPUT_PATH):
            os.remove(OUTPUT_PATH)

        # Salvar arquivo
        wb.save(OUTPUT_PATH)

        print(f"Arquivo Excel '{OUTPUT_PATH}' criado com sucesso! 🚀")

except fdb.DatabaseError as e:
    print("Erro ao conectar ou consultar o Banco de Dados!")
    print(f"Detalhes: {e}")

except PermissionError:
    print("Permissão negada ao tentar sobrescrever o arquivo Excel. Feche o arquivo se ele estiver aberto e tente novamente.")

except Exception as e:
    print(f"Erro inesperado: {e}")

finally:
    if cur is not None:
        cur.close()
    if con is not None:
        con.close()
